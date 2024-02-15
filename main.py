from web3.middleware import geth_poa_middleware

from client import Client
from data.config import private_key
from models import Arbitrum, Avalanche, Optimism, Polygon, HoleskyNetwork, TaikoNetwork
from tasks.holesky import Holesky
from tasks.taiko import Taiko
from models import TokenAmount
import openpyxl
from openpyxl import load_workbook
import time






# clientHolesky = Client(private_key=private_key, network=HoleskyNetwork)
# clientHolesky.w3.middleware_onion.inject(geth_poa_middleware, layer=0)
# holesky = Holesky(client=clientHolesky)

# clientTaiko = Client(private_key=private_key, network=TaikoNetwork)
# clientTaiko.w3.middleware_onion.inject(geth_poa_middleware, layer=0)
# taiko = Taiko(client=clientTaiko)

# block = client.w3.eth.get_block('latest')
# print(client.w3.eth.max_priority_fee)
# res = Client.get_max_priority_fee_per_gas(w3=client.w3, block=block)
# woofi = WooFi(client=client)
# amount = TokenAmount(amount=0.001)
# tx = woofi.swap_eth_to_usdc(amount=amount)
# print(res)


# holesky.mint()
# approve = holesky.bridge()
# send = holesky.bridge_horse()
# sendEth = holesky.bridge_eth()


def run_accs(accs_data: list):
    for acc in accs_data:
        print('START NEW ACCOUNT =>',acc['privatekey'])
        clientHolesky = Client(private_key=acc['privatekey'], network=HoleskyNetwork)
        clientHolesky.w3.middleware_onion.inject(geth_poa_middleware, layer=0)
        holesky = Holesky(client=clientHolesky)

        clientTaiko = Client(private_key=acc['privatekey'], network=TaikoNetwork)
        clientTaiko.w3.middleware_onion.inject(geth_poa_middleware, layer=0)
        taiko = Taiko(client=clientTaiko)

        print('START MINT =>',acc['privatekey'])
        holesky.mint()
        time.sleep(10)

        print('START APPROVING =>',acc['privatekey'])
        approve = holesky.bridge()
        time.sleep(10)

        print('START BRIDGE HORSE =>',acc['privatekey'])
        send = holesky.bridge_horse()
        time.sleep(10)

        print('START BRIDGE ETH =>',acc['privatekey'])
        sendEth = holesky.bridge_eth()
        time.sleep(10)

        print('FINISH ACCOUNT =>',acc['privatekey'])


if __name__ == '__main__':

    workbook = load_workbook(filename='accs.xlsx')
    sheet = workbook.active

    accs_data = []
    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=4):
        privatekey= row[0].value
        accs_data.append({'privatekey': privatekey})

    # try:
    run_accs(accs_data=accs_data)
    # except Exception as err:
        # print(f'Global error: {err}')

    print(f'All accs done.\n\n')

# print( "holesky aprove =>", aprove)

# send = taiko.bridge_horse()
# res = taiko.client.verif_tx(tx_hash=send)


# sendEth = taiko.pool()

# print( "holesky aprove =>", sendEth)

# print(res)


# async def read_excel(file_path):
#     workbook = openpyxl.load_workbook(file_path)
#     sheet = workbook.active
#     private_keys = []
#     proxies = []
#     for row in range(2, sheet.max_row + 1):  # Assuming the first row is headers
#         private_key = sheet['C' + str(row)].value
#         proxy = sheet['D' + str(row)].value
#         if private_key and proxy:
#             private_keys.append((private_key, row))
#             proxies.append((proxy, row))
#     return private_keys, proxies
#
#
#
# async def main():
#     private_keys, proxies = await read_excel('./accs.xlsx')
#     print(private_keys)
